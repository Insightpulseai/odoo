#!/usr/bin/env python3
"""
Odoo Import Template Generator

Generates clean CSV import templates from the embedded data dictionary.
Can be used standalone without Supabase connection.

Usage:
    python scripts/generate_odoo_template.py --model project.task --output templates/
    python scripts/generate_odoo_template.py --template finance-ppm-tasks --output templates/
    python scripts/generate_odoo_template.py --list-models
    python scripts/generate_odoo_template.py --list-templates
    python scripts/generate_odoo_template.py --all --output templates/

Options:
    --model MODEL       Generate template for a specific model
    --template SLUG     Generate template using a predefined template
    --output DIR        Output directory (default: ./templates)
    --format FORMAT     Output format: csv, xlsx, json (default: csv)
    --include-examples  Include example row in template
    --list-models       List all available models
    --list-templates    List all predefined templates
    --all               Generate all predefined templates
"""

import argparse
import csv
import json
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

# =============================================================================
# Embedded Data Dictionary (mirrors Supabase odoo_dict.fields)
# =============================================================================

@dataclass
class DictField:
    model_name: str
    field_name: str
    label: str
    field_type: str
    required: bool
    is_key: bool
    relation_model: Optional[str]
    import_column: str
    description: Optional[str]
    example_value: Optional[str]
    default_value: Optional[str]
    domain: str
    sequence: int


@dataclass
class Template:
    slug: str
    name: str
    description: str
    model_name: str
    field_names: list[str]
    domain: str


# Embedded field definitions (canonical source of truth)
FIELDS: list[DictField] = [
    # =========================================================================
    # PROJECT.PROJECT Fields
    # =========================================================================
    DictField(
        model_name="project.project",
        field_name="x_external_ref",
        label="External Reference",
        field_type="char",
        required=True,
        is_key=True,
        relation_model=None,
        import_column="External ID",
        description="Stable key for this project across environments. Used for idempotent imports.",
        example_value="FIN_CLOSE_TBWA",
        default_value=None,
        domain="project",
        sequence=10,
    ),
    DictField(
        model_name="project.project",
        field_name="name",
        label="Project Name",
        field_type="char",
        required=True,
        is_key=False,
        relation_model=None,
        import_column="Name",
        description="Human-friendly project name displayed in UI.",
        example_value="Month-end Closing & Tax Filing",
        default_value=None,
        domain="project",
        sequence=20,
    ),
    DictField(
        model_name="project.project",
        field_name="description",
        label="Description",
        field_type="html",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Description",
        description="Rich text description of the project scope and objectives.",
        example_value="<p>Finance PPM for month-end close and BIR tax compliance</p>",
        default_value=None,
        domain="project",
        sequence=30,
    ),
    DictField(
        model_name="project.project",
        field_name="partner_id",
        label="Customer",
        field_type="many2one",
        required=False,
        is_key=False,
        relation_model="res.partner",
        import_column="Customer",
        description="Legal entity or client owning this project.",
        example_value="TBWA SMP",
        default_value=None,
        domain="project",
        sequence=40,
    ),
    DictField(
        model_name="project.project",
        field_name="company_id",
        label="Company",
        field_type="many2one",
        required=True,
        is_key=False,
        relation_model="res.company",
        import_column="Company",
        description="Odoo company owning this project. Required for multi-company finance.",
        example_value="TBWA\\SMP",
        default_value=None,
        domain="project",
        sequence=50,
    ),
    DictField(
        model_name="project.project",
        field_name="user_id",
        label="Project Manager",
        field_type="many2one",
        required=False,
        is_key=False,
        relation_model="res.users",
        import_column="Project Manager",
        description="User responsible for overall project delivery.",
        example_value="Rey Meran",
        default_value=None,
        domain="project",
        sequence=60,
    ),
    DictField(
        model_name="project.project",
        field_name="privacy_visibility",
        label="Visibility",
        field_type="selection",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Visibility",
        description="Who can see this project: portal, employees, or followers.",
        example_value="employees",
        default_value="employees",
        domain="project",
        sequence=70,
    ),
    DictField(
        model_name="project.project",
        field_name="allow_task_dependencies",
        label="Allow Dependencies",
        field_type="boolean",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Allow Task Dependencies",
        description="Enable task dependency tracking (CE feature).",
        example_value="True",
        default_value="True",
        domain="project",
        sequence=80,
    ),
    DictField(
        model_name="project.project",
        field_name="date_start",
        label="Start Date",
        field_type="date",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Start Date",
        description="Planned project start date.",
        example_value="2026-01-01",
        default_value=None,
        domain="project",
        sequence=90,
    ),
    DictField(
        model_name="project.project",
        field_name="date",
        label="End Date",
        field_type="date",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Deadline",
        description="Planned project end date or deadline.",
        example_value="2026-12-31",
        default_value=None,
        domain="project",
        sequence=100,
    ),
    DictField(
        model_name="project.project",
        field_name="active",
        label="Active",
        field_type="boolean",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Active",
        description="Whether the project is active.",
        example_value="True",
        default_value="True",
        domain="project",
        sequence=110,
    ),

    # =========================================================================
    # PROJECT.TASK Fields
    # =========================================================================
    DictField(
        model_name="project.task",
        field_name="x_external_ref",
        label="External Reference",
        field_type="char",
        required=True,
        is_key=True,
        relation_model=None,
        import_column="External ID",
        description="Stable key for this task. Used for idempotent imports and dependency references.",
        example_value="FIN_CLOSE_DAY3_TRIALBAL",
        default_value=None,
        domain="finance",
        sequence=10,
    ),
    DictField(
        model_name="project.task",
        field_name="name",
        label="Task Name",
        field_type="char",
        required=True,
        is_key=False,
        relation_model=None,
        import_column="Name of the Tasks?",
        description="Name of the finance/closing task.",
        example_value="Post Trial Balance & Adjustments",
        default_value=None,
        domain="finance",
        sequence=20,
    ),
    DictField(
        model_name="project.task",
        field_name="description",
        label="Description",
        field_type="html",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Description",
        description="Rich text description of what this task entails.",
        example_value="<p>Run trial balance, post adjusting entries</p>",
        default_value=None,
        domain="finance",
        sequence=30,
    ),
    DictField(
        model_name="project.task",
        field_name="project_id",
        label="Project",
        field_type="many2one",
        required=True,
        is_key=False,
        relation_model="project.project",
        import_column="Project",
        description="Parent project containing this task.",
        example_value="Month-end Closing & Tax Filing",
        default_value=None,
        domain="finance",
        sequence=40,
    ),
    DictField(
        model_name="project.task",
        field_name="company_id",
        label="Company",
        field_type="many2one",
        required=True,
        is_key=False,
        relation_model="res.company",
        import_column="Company",
        description="Odoo company for multi-company finance setup.",
        example_value="TBWA\\SMP",
        default_value=None,
        domain="finance",
        sequence=50,
    ),
    DictField(
        model_name="project.task",
        field_name="user_ids",
        label="Assigned to",
        field_type="many2many",
        required=False,
        is_key=False,
        relation_model="res.users",
        import_column="Assigned to",
        description="One or more assignees (Finance Director, Sr Finance Manager, etc.).",
        example_value="Rey Meran",
        default_value=None,
        domain="finance",
        sequence=60,
    ),
    DictField(
        model_name="project.task",
        field_name="stage_id",
        label="Stage",
        field_type="many2one",
        required=False,
        is_key=False,
        relation_model="project.task.type",
        import_column="Stage",
        description="Current workflow stage of the task.",
        example_value="To Do",
        default_value="To Do",
        domain="finance",
        sequence=70,
    ),
    DictField(
        model_name="project.task",
        field_name="planned_hours",
        label="Allocated Time",
        field_type="float",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Allocated Time",
        description="Planned effort in hours for this task per cycle.",
        example_value="4.0",
        default_value="0",
        domain="finance",
        sequence=80,
    ),
    DictField(
        model_name="project.task",
        field_name="date_deadline",
        label="Deadline",
        field_type="date",
        required=True,
        is_key=False,
        relation_model=None,
        import_column="Planned Date",
        description="Operational due date (e.g., Day 3, Day 5, BIR statutory cutoffs).",
        example_value="2026-01-05",
        default_value=None,
        domain="finance",
        sequence=90,
    ),
    DictField(
        model_name="project.task",
        field_name="tag_ids",
        label="Tags",
        field_type="many2many",
        required=False,
        is_key=False,
        relation_model="project.tags",
        import_column="Tags",
        description="Classification tags: Month-End, Tax Filing, Risk, Must-Do, etc.",
        example_value="Month-End,Tax Filing",
        default_value=None,
        domain="finance",
        sequence=100,
    ),
    DictField(
        model_name="project.task",
        field_name="priority",
        label="Priority",
        field_type="selection",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Priority",
        description="Task priority: 0=Normal, 1=Important.",
        example_value="1",
        default_value="0",
        domain="finance",
        sequence=110,
    ),
    DictField(
        model_name="project.task",
        field_name="depend_on_ids",
        label="Dependencies",
        field_type="many2many",
        required=False,
        is_key=False,
        relation_model="project.task",
        import_column="Depends on",
        description="Tasks that must complete before this task can start.",
        example_value="FIN_CLOSE_DAY1_CUTOFF",
        default_value=None,
        domain="finance",
        sequence=120,
    ),
    DictField(
        model_name="project.task",
        field_name="sequence",
        label="Sequence",
        field_type="integer",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Sequence",
        description="Order of task in project view. Lower = earlier.",
        example_value="10",
        default_value="10",
        domain="finance",
        sequence=130,
    ),
    DictField(
        model_name="project.task",
        field_name="parent_id",
        label="Parent Task",
        field_type="many2one",
        required=False,
        is_key=False,
        relation_model="project.task",
        import_column="Parent Task",
        description="Parent task if this is a subtask.",
        example_value=None,
        default_value=None,
        domain="finance",
        sequence=140,
    ),
    DictField(
        model_name="project.task",
        field_name="active",
        label="Active",
        field_type="boolean",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Active",
        description="Whether the task is active.",
        example_value="True",
        default_value="True",
        domain="finance",
        sequence=150,
    ),

    # =========================================================================
    # PROJECT.TAGS Fields
    # =========================================================================
    DictField(
        model_name="project.tags",
        field_name="x_external_ref",
        label="External Reference",
        field_type="char",
        required=False,
        is_key=True,
        relation_model=None,
        import_column="External ID",
        description="Stable key for this tag.",
        example_value="TAG_MONTH_END",
        default_value=None,
        domain="project",
        sequence=10,
    ),
    DictField(
        model_name="project.tags",
        field_name="name",
        label="Tag Name",
        field_type="char",
        required=True,
        is_key=False,
        relation_model=None,
        import_column="Name",
        description="Display name of the tag.",
        example_value="Month-End",
        default_value=None,
        domain="project",
        sequence=20,
    ),
    DictField(
        model_name="project.tags",
        field_name="color",
        label="Color",
        field_type="integer",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Color Index",
        description="Color index (0-11) for visual differentiation.",
        example_value="2",
        default_value="0",
        domain="project",
        sequence=30,
    ),

    # =========================================================================
    # PROJECT.TASK.TYPE (Stages) Fields
    # =========================================================================
    DictField(
        model_name="project.task.type",
        field_name="x_external_ref",
        label="External Reference",
        field_type="char",
        required=False,
        is_key=True,
        relation_model=None,
        import_column="External ID",
        description="Stable key for this stage.",
        example_value="STAGE_TODO",
        default_value=None,
        domain="project",
        sequence=10,
    ),
    DictField(
        model_name="project.task.type",
        field_name="name",
        label="Stage Name",
        field_type="char",
        required=True,
        is_key=False,
        relation_model=None,
        import_column="Name",
        description="Display name of the stage.",
        example_value="To Do",
        default_value=None,
        domain="project",
        sequence=20,
    ),
    DictField(
        model_name="project.task.type",
        field_name="sequence",
        label="Sequence",
        field_type="integer",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Sequence",
        description="Order of stage in kanban. Lower = earlier.",
        example_value="1",
        default_value="1",
        domain="project",
        sequence=30,
    ),
    DictField(
        model_name="project.task.type",
        field_name="fold",
        label="Folded",
        field_type="boolean",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Folded in Kanban",
        description="Whether this stage is collapsed by default in kanban.",
        example_value="False",
        default_value="False",
        domain="project",
        sequence=40,
    ),
    DictField(
        model_name="project.task.type",
        field_name="project_ids",
        label="Projects",
        field_type="many2many",
        required=False,
        is_key=False,
        relation_model="project.project",
        import_column="Projects",
        description="Projects that use this stage.",
        example_value="Month-end Closing & Tax Filing",
        default_value=None,
        domain="project",
        sequence=50,
    ),

    # =========================================================================
    # HR.EMPLOYEE Fields
    # =========================================================================
    DictField(
        model_name="hr.employee",
        field_name="x_external_ref",
        label="External Reference",
        field_type="char",
        required=False,
        is_key=True,
        relation_model=None,
        import_column="External ID",
        description="Stable key for this employee.",
        example_value="EMP_FINANCE_DIR",
        default_value=None,
        domain="hr",
        sequence=10,
    ),
    DictField(
        model_name="hr.employee",
        field_name="name",
        label="Employee Name",
        field_type="char",
        required=True,
        is_key=False,
        relation_model=None,
        import_column="Name",
        description="Full name of the employee.",
        example_value="Rey Meran",
        default_value=None,
        domain="hr",
        sequence=20,
    ),
    DictField(
        model_name="hr.employee",
        field_name="work_email",
        label="Work Email",
        field_type="char",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Work Email",
        description="Corporate email address.",
        example_value="rey.meran@tbwa.com.ph",
        default_value=None,
        domain="hr",
        sequence=30,
    ),
    DictField(
        model_name="hr.employee",
        field_name="job_title",
        label="Job Title",
        field_type="char",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Job Title",
        description="Current job title/position.",
        example_value="Finance Director",
        default_value=None,
        domain="hr",
        sequence=40,
    ),
    DictField(
        model_name="hr.employee",
        field_name="department_id",
        label="Department",
        field_type="many2one",
        required=False,
        is_key=False,
        relation_model="hr.department",
        import_column="Department",
        description="Department this employee belongs to.",
        example_value="Finance",
        default_value=None,
        domain="hr",
        sequence=50,
    ),
    DictField(
        model_name="hr.employee",
        field_name="company_id",
        label="Company",
        field_type="many2one",
        required=True,
        is_key=False,
        relation_model="res.company",
        import_column="Company",
        description="Company this employee works for.",
        example_value="TBWA\\SMP",
        default_value=None,
        domain="hr",
        sequence=60,
    ),
    DictField(
        model_name="hr.employee",
        field_name="user_id",
        label="Related User",
        field_type="many2one",
        required=False,
        is_key=False,
        relation_model="res.users",
        import_column="Related User",
        description="Odoo user account linked to this employee.",
        example_value="rey.meran@tbwa.com.ph",
        default_value=None,
        domain="hr",
        sequence=70,
    ),
    DictField(
        model_name="hr.employee",
        field_name="active",
        label="Active",
        field_type="boolean",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Active",
        description="Whether the employee is currently active.",
        example_value="True",
        default_value="True",
        domain="hr",
        sequence=80,
    ),

    # =========================================================================
    # ACCOUNT.ANALYTIC.ACCOUNT Fields
    # =========================================================================
    DictField(
        model_name="account.analytic.account",
        field_name="x_external_ref",
        label="External Reference",
        field_type="char",
        required=False,
        is_key=True,
        relation_model=None,
        import_column="External ID",
        description="Stable key for this analytic account.",
        example_value="AA_FIN_CLOSE_2026",
        default_value=None,
        domain="finance",
        sequence=10,
    ),
    DictField(
        model_name="account.analytic.account",
        field_name="name",
        label="Name",
        field_type="char",
        required=True,
        is_key=False,
        relation_model=None,
        import_column="Name",
        description="Name of the analytic account.",
        example_value="Finance Close 2026",
        default_value=None,
        domain="finance",
        sequence=20,
    ),
    DictField(
        model_name="account.analytic.account",
        field_name="code",
        label="Reference",
        field_type="char",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Reference",
        description="Short code/reference for the analytic account.",
        example_value="FIN-CLOSE-26",
        default_value=None,
        domain="finance",
        sequence=30,
    ),
    DictField(
        model_name="account.analytic.account",
        field_name="plan_id",
        label="Analytic Plan",
        field_type="many2one",
        required=True,
        is_key=False,
        relation_model="account.analytic.plan",
        import_column="Analytic Plan",
        description="Plan this account belongs to.",
        example_value="Projects",
        default_value=None,
        domain="finance",
        sequence=40,
    ),
    DictField(
        model_name="account.analytic.account",
        field_name="company_id",
        label="Company",
        field_type="many2one",
        required=False,
        is_key=False,
        relation_model="res.company",
        import_column="Company",
        description="Company owning this analytic account.",
        example_value="TBWA\\SMP",
        default_value=None,
        domain="finance",
        sequence=50,
    ),
    DictField(
        model_name="account.analytic.account",
        field_name="active",
        label="Active",
        field_type="boolean",
        required=False,
        is_key=False,
        relation_model=None,
        import_column="Active",
        description="Whether the analytic account is active.",
        example_value="True",
        default_value="True",
        domain="finance",
        sequence=60,
    ),
]

# Predefined templates
TEMPLATES: list[Template] = [
    Template(
        slug="finance-ppm-tasks",
        name="Finance PPM Tasks",
        description="Full task import for month-end closing and tax filing",
        model_name="project.task",
        field_names=[
            "x_external_ref", "name", "description", "project_id", "company_id",
            "user_ids", "stage_id", "planned_hours", "date_deadline", "tag_ids",
            "priority", "depend_on_ids", "sequence"
        ],
        domain="finance",
    ),
    Template(
        slug="finance-ppm-tasks-minimal",
        name="Finance PPM Tasks (Minimal)",
        description="Minimal task import with required fields only",
        model_name="project.task",
        field_names=["x_external_ref", "name", "project_id", "company_id", "date_deadline"],
        domain="finance",
    ),
    Template(
        slug="finance-ppm-projects",
        name="Finance PPM Projects",
        description="Project import for finance programs",
        model_name="project.project",
        field_names=["x_external_ref", "name", "description", "company_id", "user_id", "date_start", "date"],
        domain="finance",
    ),
    Template(
        slug="project-stages",
        name="Project Stages",
        description="Stage definitions for project kanban",
        model_name="project.task.type",
        field_names=["x_external_ref", "name", "sequence", "fold"],
        domain="project",
    ),
    Template(
        slug="project-tags",
        name="Project Tags",
        description="Tag definitions for task categorization",
        model_name="project.tags",
        field_names=["x_external_ref", "name", "color"],
        domain="project",
    ),
    Template(
        slug="hr-employees",
        name="HR Employees",
        description="Employee master data for assignment references",
        model_name="hr.employee",
        field_names=["x_external_ref", "name", "work_email", "job_title", "department_id", "company_id"],
        domain="hr",
    ),
    Template(
        slug="analytic-accounts",
        name="Analytic Accounts",
        description="Analytic accounts for project cost tracking",
        model_name="account.analytic.account",
        field_names=["x_external_ref", "name", "code", "plan_id", "company_id"],
        domain="finance",
    ),
]


# =============================================================================
# Helper Functions
# =============================================================================

def get_fields_for_model(model_name: str) -> list[DictField]:
    """Get all fields for a model, sorted by sequence."""
    fields = [f for f in FIELDS if f.model_name == model_name]
    return sorted(fields, key=lambda f: (f.sequence, f.field_name))


def get_fields_for_template(template: Template) -> list[DictField]:
    """Get fields for a template in the specified order."""
    field_map = {f.field_name: f for f in FIELDS if f.model_name == template.model_name}
    return [field_map[name] for name in template.field_names if name in field_map]


def get_available_models() -> list[str]:
    """Get list of all models in the dictionary."""
    return sorted(set(f.model_name for f in FIELDS))


def get_template_by_slug(slug: str) -> Optional[Template]:
    """Get template by slug."""
    for t in TEMPLATES:
        if t.slug == slug:
            return t
    return None


# =============================================================================
# Output Generators
# =============================================================================

def generate_csv(fields: list[DictField], include_examples: bool = False) -> str:
    """Generate CSV content from fields."""
    import io

    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    headers = [f.import_column for f in fields]
    writer.writerow(headers)

    # Example row if requested
    if include_examples:
        examples = [f.example_value or "" for f in fields]
        writer.writerow(examples)

    return output.getvalue()


def generate_json(fields: list[DictField], model_name: str, template_slug: Optional[str] = None) -> str:
    """Generate JSON metadata from fields."""
    data = {
        "model": model_name,
        "template": template_slug,
        "fields": [
            {
                "field_name": f.field_name,
                "import_column": f.import_column,
                "label": f.label,
                "field_type": f.field_type,
                "required": f.required,
                "is_key": f.is_key,
                "relation_model": f.relation_model,
                "description": f.description,
                "example_value": f.example_value,
                "default_value": f.default_value,
            }
            for f in fields
        ],
    }
    return json.dumps(data, indent=2)


def generate_xlsx(fields: list[DictField], include_examples: bool = False) -> bytes:
    """Generate XLSX content from fields."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for XLSX output. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Header row
    for col, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col, value=field.import_column)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(15, len(field.import_column) + 2)

    # Example row if requested
    if include_examples:
        for col, field in enumerate(fields, start=1):
            ws.cell(row=2, column=col, value=field.example_value or "")

    # Save to bytes
    import io
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# =============================================================================
# Main CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate Odoo import templates from data dictionary",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --model project.task --output templates/
  %(prog)s --template finance-ppm-tasks --include-examples
  %(prog)s --list-templates
  %(prog)s --all --output templates/
        """,
    )

    parser.add_argument("--model", "-m", help="Model name (e.g., project.task)")
    parser.add_argument("--template", "-t", help="Template slug (e.g., finance-ppm-tasks)")
    parser.add_argument("--output", "-o", default=".", help="Output directory (default: .)")
    parser.add_argument("--format", "-f", choices=["csv", "xlsx", "json"], default="csv", help="Output format")
    parser.add_argument("--include-examples", "-e", action="store_true", help="Include example row")
    parser.add_argument("--list-models", action="store_true", help="List available models")
    parser.add_argument("--list-templates", action="store_true", help="List available templates")
    parser.add_argument("--all", "-a", action="store_true", help="Generate all templates")

    args = parser.parse_args()

    # Handle list commands
    if args.list_models:
        print("Available models:")
        for model in get_available_models():
            field_count = len(get_fields_for_model(model))
            print(f"  {model} ({field_count} fields)")
        return 0

    if args.list_templates:
        print("Available templates:")
        for t in TEMPLATES:
            print(f"  {t.slug}")
            print(f"    Model: {t.model_name}")
            print(f"    Fields: {len(t.field_names)}")
            print(f"    Description: {t.description}")
            print()
        return 0

    # Ensure output directory exists
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate all templates
    if args.all:
        for template in TEMPLATES:
            fields = get_fields_for_template(template)
            generate_and_save(fields, template.model_name, template.slug, output_dir, args.format, args.include_examples)
        print(f"Generated {len(TEMPLATES)} templates in {output_dir}")
        return 0

    # Generate single template by slug
    if args.template:
        template = get_template_by_slug(args.template)
        if not template:
            print(f"Error: Template '{args.template}' not found", file=sys.stderr)
            print("Use --list-templates to see available templates", file=sys.stderr)
            return 1
        fields = get_fields_for_template(template)
        generate_and_save(fields, template.model_name, template.slug, output_dir, args.format, args.include_examples)
        return 0

    # Generate by model name
    if args.model:
        fields = get_fields_for_model(args.model)
        if not fields:
            print(f"Error: Model '{args.model}' not found in dictionary", file=sys.stderr)
            print("Use --list-models to see available models", file=sys.stderr)
            return 1
        generate_and_save(fields, args.model, None, output_dir, args.format, args.include_examples)
        return 0

    # No action specified
    parser.print_help()
    return 1


def generate_and_save(
    fields: list[DictField],
    model_name: str,
    template_slug: Optional[str],
    output_dir: Path,
    fmt: str,
    include_examples: bool,
):
    """Generate template and save to file."""
    filename_base = template_slug or model_name.replace(".", "_")

    if fmt == "csv":
        content = generate_csv(fields, include_examples)
        filepath = output_dir / f"{filename_base}.csv"
        filepath.write_text(content)
    elif fmt == "json":
        content = generate_json(fields, model_name, template_slug)
        filepath = output_dir / f"{filename_base}.json"
        filepath.write_text(content)
    elif fmt == "xlsx":
        content = generate_xlsx(fields, include_examples)
        filepath = output_dir / f"{filename_base}.xlsx"
        filepath.write_bytes(content)

    print(f"Generated: {filepath}")


if __name__ == "__main__":
    sys.exit(main())
