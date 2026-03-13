#!/usr/bin/env python3
"""Export canonical seed CSVs from the Finance PPM workbook.

Canonical source:
    config/finance/Month-end Closing Task and Tax Filing (7).xlsx
    (or user-specified path via --workbook)

Outputs (written to config/finance/):
    - month_end_tasks_v2.csv      (36 closing tasks with durations + dependencies)
    - bir_filing_tasks_v2.csv     (22 filing tasks + 5 process steps)
    - finance_odoo_import.csv      (168 Odoo-importable task rows)
    - supabase_logframe_finance.csv (168 logframe rows)

Date correction: BIR filing dates are normalised to 2026 context year.
"""

import argparse
import csv
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CONTEXT_YEAR = 2026  # All BIR deadlines target this fiscal year

# Task codes and IDs follow the v2 convention from the original seed data
MONTH_END_TASK_CODES = [
    "PAYROLL_001", "TAX_PROV_001", "RENT_001", "ACCRUAL_001", "ACCRUAL_002",
    "PRIOR_001", "AMORT_001", "CORP_ACC_001", "INSURE_001", "TREASURY_001",
    "PRIOR_002", "REGIONAL_001", "BILLING_001", "WIP_OOP_001", "AMORT_002",
    "AMORT_003", "AMORT_004", "AMORT_005", "AR_WC_001", "CA_LIQ_001",
    "AP_WC_001", "OOP_001", "ASSET_001", "RECLASS_001", "VAT_001",
    "ACC_ASSET_001", "ACC_ASSET_002", "AP_001", "CA_LIQ_002", "ACC_ASSET_003",
    "EXP_REC_001", "VAT_RPT_001", "JOB_XFER_001", "JOB_XFER_002",
    "ACCRUALS_001", "WIP_001",
]

# Dependencies (task_code → list of prerequisite task_codes)
MONTH_END_DEPS = {
    "TAX_PROV_001": ["PAYROLL_001"],
    "REGIONAL_001": ["BILLING_001"],
    "WIP_OOP_001": ["BILLING_001"],
    "AR_WC_001": ["BILLING_001"],
    "CA_LIQ_001": ["ACCRUAL_002"],
    "AP_WC_001": ["AP_001"],
    "OOP_001": ["WIP_OOP_001"],
    "RECLASS_001": ["BILLING_001"],
    "VAT_001": ["PAYROLL_001", "BILLING_001", "ACCRUAL_001"],
    "AP_001": ["ACC_ASSET_001", "ACC_ASSET_002"],
    "VAT_RPT_001": ["VAT_001"],
    "JOB_XFER_001": ["BILLING_001"],
    "JOB_XFER_002": ["RECLASS_001"],
    "ACCRUALS_001": ["BILLING_001"],
    "WIP_001": ["WIP_OOP_001", "BILLING_001"],
}


def parse_duration(val):
    """Convert '1 Day' / '0.5 Day' strings to float days."""
    if val is None:
        return 0.0
    s = str(val).strip().lower().replace("days", "").replace("day", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def fix_bir_year(dt, context_year=CONTEXT_YEAR):
    """Correct BIR dates: if year < context_year, bump to context_year."""
    if dt is None:
        return None
    if isinstance(dt, str):
        return dt  # pass-through for text periods like 'Q4 2025'
    if isinstance(dt, datetime) and dt.year < context_year:
        try:
            return dt.replace(year=context_year)
        except ValueError:
            # Handle Feb 29 → Feb 28 edge case
            return dt.replace(year=context_year, day=28)
    return dt


def fmt_date(dt):
    """Format datetime → YYYY-MM-DD string."""
    if dt is None:
        return ""
    if isinstance(dt, str):
        return dt
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d")
    return str(dt)


def fmt_period(dt):
    """Format period: datetime → 'Mon YYYY', string → pass-through."""
    if dt is None:
        return ""
    if isinstance(dt, str):
        return dt
    if isinstance(dt, datetime):
        return dt.strftime("%b %Y")
    return str(dt)


# ---------------------------------------------------------------------------
# Exporters
# ---------------------------------------------------------------------------

def export_month_end(wb, out_dir):
    """Export Closing Task sheet → month_end_tasks_v2.csv."""
    ws = wb["Closing Task"]
    rows = []
    current_employee = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row)
        # Stop at the employee reference section
        if vals[0] == "EMPLOYEE CODE REFERENCE":
            break
        if not any(v is not None for v in vals[:8]):
            continue

        # Carry forward employee code
        if vals[0] is not None:
            current_employee = vals[0]
        employee_code = current_employee

        category = vals[1] if vals[1] else (rows[-1]["category"] if rows else "")
        name = vals[2]
        reviewed_by = vals[3]
        approved_by = vals[4]
        prep_days = parse_duration(vals[5])
        review_days = parse_duration(vals[6])
        approval_days = parse_duration(vals[7])
        total_days = prep_days + review_days + approval_days
        assignee_name = vals[8] if len(vals) > 8 else ""
        assignee_email = vals[9] if len(vals) > 9 else ""

        if name is None:
            continue

        idx = len(rows)
        task_code = MONTH_END_TASK_CODES[idx] if idx < len(MONTH_END_TASK_CODES) else f"TASK_{idx+1:03d}"
        task_id = f"task_close_{idx+1:02d}"
        depends = ",".join(MONTH_END_DEPS.get(task_code, []))

        rows.append({
            "task_code": task_code,
            "task_id": task_id,
            "name": name,
            "category": category,
            "employee_code": employee_code,
            "reviewed_by": reviewed_by,
            "approved_by": approved_by,
            "planned_hours": total_days * 8,  # 8h per day
            "prep_days": prep_days,
            "review_days": review_days,
            "approval_days": approval_days,
            "total_days": total_days,
            "depends_on": depends,
            "assignee_name": assignee_name,
            "assignee_email": assignee_email,
        })

    outpath = os.path.join(out_dir, "month_end_tasks_v2.csv")
    fieldnames = [
        "task_code", "task_id", "name", "category", "employee_code",
        "reviewed_by", "approved_by", "planned_hours", "prep_days",
        "review_days", "approval_days", "total_days", "depends_on",
        "assignee_name", "assignee_email",
    ]
    with open(outpath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"  month_end_tasks_v2.csv: {len(rows)} tasks")
    return len(rows)


def export_bir_filing(wb, out_dir):
    """Export Tax Filing sheet → bir_filing_tasks_v2.csv."""
    ws = wb["Tax Filing"]
    filing_rows = []
    process_rows = []
    in_process_section = False
    seq = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not any(v is not None for v in vals):
            continue

        # Detect process steps section
        if vals[0] == "Step":
            in_process_section = True
            continue

        if in_process_section:
            step_name = vals[0]
            detail = vals[1]
            if step_name and detail:
                seq += 1
                code = f"TAX_STEP_{seq:03d}"
                process_rows.append({
                    "task_code": code,
                    "bir_form": step_name,
                    "period": detail,
                    "deadline": "",
                    "prep_date": "",
                    "review_date": "",
                    "approval_date": "",
                    "prep_days": "",
                    "review_days": "",
                    "approval_days": "",
                    "total_days": "",
                    "depends_on": "",
                    "is_metadata": "true",
                })
            continue

        # Filing task row
        bir_form = vals[0]
        period_raw = vals[1]
        deadline_raw = vals[2]
        prep_raw = vals[3]
        review_raw = vals[4]
        approval_raw = vals[5]

        if bir_form is None:
            continue

        # Fix dates to 2026 context year
        deadline = fix_bir_year(deadline_raw)
        prep = fix_bir_year(prep_raw)
        review = fix_bir_year(review_raw)
        approval = fix_bir_year(approval_raw)

        # Calculate durations
        prep_days = ""
        review_days = ""
        approval_days = ""
        total_days = ""
        if isinstance(prep, datetime) and isinstance(review, datetime) and isinstance(approval, datetime) and isinstance(deadline, datetime):
            p = (review - prep).days
            r = (approval - review).days
            a = (deadline - approval).days
            prep_days = max(p, 1)
            review_days = max(r, 1)
            approval_days = max(a, 1)
            total_days = prep_days + review_days + approval_days

        period_str = fmt_period(period_raw)
        seq += 1
        # Generate task code from BIR form
        form_slug = bir_form.replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "").replace("-", "_").upper()
        code = f"TAX_{form_slug}_{seq:03d}"

        filing_rows.append({
            "task_code": code,
            "bir_form": bir_form,
            "period": period_str,
            "deadline": fmt_date(deadline),
            "prep_date": fmt_date(prep),
            "review_date": fmt_date(review),
            "approval_date": fmt_date(approval),
            "prep_days": prep_days,
            "review_days": review_days,
            "approval_days": approval_days,
            "total_days": total_days,
            "depends_on": "",
            "is_metadata": "false",
        })

    all_rows = filing_rows + process_rows
    outpath = os.path.join(out_dir, "bir_filing_tasks_v2.csv")
    fieldnames = [
        "task_code", "bir_form", "period", "deadline", "prep_date",
        "review_date", "approval_date", "prep_days", "review_days",
        "approval_days", "total_days", "depends_on", "is_metadata",
    ]
    with open(outpath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"  bir_filing_tasks_v2.csv: {len(filing_rows)} filings + {len(process_rows)} process steps = {len(all_rows)} rows")
    return len(filing_rows), len(process_rows)


def export_odoo_import(wb, out_dir):
    """Export Odoo Import (project.task) → odoo_import_project_task.csv."""
    ws = wb["Odoo Import (project.task)"]
    headers = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not any(v is not None for v in vals):
            continue
        row_dict = dict(zip(headers, vals))
        # Normalise deadline to string
        if row_dict.get("Deadline") and isinstance(row_dict["Deadline"], datetime):
            row_dict["Deadline"] = row_dict["Deadline"].strftime("%Y-%m-%d")
        rows.append(row_dict)

    outpath = os.path.join(out_dir, "finance_odoo_import.csv")
    with open(outpath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    print(f"  finance_odoo_import.csv: {len(rows)} tasks")
    return len(rows)


def export_supabase_logframe(wb, out_dir):
    """Export Supabase Logframe CSV → supabase_logframe_finance.csv."""
    ws = wb["Supabase Logframe CSV"]
    headers = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not any(v is not None for v in vals):
            continue
        row_dict = dict(zip(headers, vals))
        # Normalise deadline/date fields
        if row_dict.get("deadline") and isinstance(row_dict["deadline"], datetime):
            row_dict["deadline"] = row_dict["deadline"].strftime("%Y-%m-%d")
        rows.append(row_dict)

    outpath = os.path.join(out_dir, "supabase_logframe_finance.csv")
    with open(outpath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    print(f"  supabase_logframe_finance.csv: {len(rows)} rows")
    return len(rows)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate(counts):
    """Validate export counts against expected values."""
    errors = []
    me_count = counts.get("month_end", 0)
    bir_filings = counts.get("bir_filings", 0)
    bir_steps = counts.get("bir_steps", 0)
    odoo_count = counts.get("odoo_import", 0)
    supa_count = counts.get("supabase_logframe", 0)

    if me_count != 36:
        errors.append(f"month_end: expected 36 tasks, got {me_count}")
    if bir_filings < 22:
        errors.append(f"bir_filing: expected >=22 filing tasks, got {bir_filings}")
    if bir_steps < 4:
        errors.append(f"bir_filing: expected >=4 process steps, got {bir_steps}")
    if odoo_count < 100:
        errors.append(f"odoo_import: expected >=100 rows, got {odoo_count}")
    if supa_count < 100:
        errors.append(f"supabase_logframe: expected >=100 rows, got {supa_count}")

    return errors


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        default=None,
        help="Path to the xlsx workbook (default: auto-detect in config/finance/)",
    )
    parser.add_argument(
        "--out-dir",
        default=None,
        help="Output directory (default: config/finance/ relative to repo root)",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Only validate existing CSVs, don't export",
    )
    args = parser.parse_args()

    # Resolve repo root
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent.parent
    out_dir = args.out_dir or str(repo_root / "config" / "finance")

    if args.validate_only:
        print("Validating existing CSVs...")
        counts = {}
        for name, key in [
            ("month_end_tasks_v2.csv", "month_end"),
            ("bir_filing_tasks_v2.csv", None),
            ("finance_odoo_import.csv", "odoo_import"),
            ("supabase_logframe_finance.csv", "supabase_logframe"),
        ]:
            path = os.path.join(out_dir, name)
            if not os.path.exists(path):
                print(f"  MISSING: {name}")
                continue
            with open(path) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                if key:
                    counts[key] = len(rows)
                elif name.startswith("bir"):
                    filings = [r for r in rows if r.get("is_metadata", "false") == "false"]
                    steps = [r for r in rows if r.get("is_metadata", "false") == "true"]
                    counts["bir_filings"] = len(filings)
                    counts["bir_steps"] = len(steps)
                print(f"  {name}: {len(rows)} rows")

        errors = validate(counts)
        if errors:
            print("\nVALIDATION FAILED:")
            for e in errors:
                print(f"  - {e}")
            sys.exit(1)
        print("\nAll validations passed.")
        sys.exit(0)

    # Find workbook
    workbook_path = args.workbook
    if not workbook_path:
        # Auto-detect xlsx in config/finance/
        for f in os.listdir(out_dir):
            if f.endswith(".xlsx"):
                workbook_path = os.path.join(out_dir, f)
                break
    if not workbook_path or not os.path.exists(workbook_path):
        print("ERROR: No workbook found. Specify --workbook PATH", file=sys.stderr)
        sys.exit(1)

    print(f"Source: {workbook_path}")
    print(f"Output: {out_dir}")
    print()

    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    counts = {}
    counts["month_end"] = export_month_end(wb, out_dir)
    bir_f, bir_s = export_bir_filing(wb, out_dir)
    counts["bir_filings"] = bir_f
    counts["bir_steps"] = bir_s
    counts["odoo_import"] = export_odoo_import(wb, out_dir)
    counts["supabase_logframe"] = export_supabase_logframe(wb, out_dir)

    print()
    errors = validate(counts)
    if errors:
        print("VALIDATION WARNINGS:")
        for e in errors:
            print(f"  - {e}")
    else:
        print("All validations passed.")


if __name__ == "__main__":
    main()
