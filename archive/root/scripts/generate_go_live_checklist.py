#!/usr/bin/env python3
import csv
import os
import sys

# Ensure openpyxl is installed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl not installed. Please run: pip install openpyxl")
    sys.exit(1)

OUTPUT_DIR = "artifacts"
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

out_xlsx = os.path.join(OUTPUT_DIR, "TBWA_OMC_Odoo18_GoLive_Checklist.xlsx")
out_csv = os.path.join(OUTPUT_DIR, "TBWA_OMC_Odoo18_GoLive_Checklist.csv")

# Columns
cols = [
    "Track",
    "Section",
    "Step ID",
    "Task",
    "Owner Role",
    "Owner Login/Email (optional)",
    "Due Date",
    "Gate (Y/N)",
    "Evidence/Link",
    "Status",
    "Notes",
]

# Status options
status_opts = ["Not Started", "In Progress", "Blocked", "Done"]
gate_opts = ["Y", "N"]

# Checklist rows (tailored)
rows = []


def add(track, section, step, task, role="", gate="N"):
    rows.append(
        {
            "Track": track,
            "Section": section,
            "Step ID": step,
            "Task": task,
            "Owner Role": role,
            "Owner Login/Email (optional)": "",
            "Due Date": "",
            "Gate (Y/N)": gate,
            "Evidence/Link": "",
            "Status": "Not Started",
            "Notes": "",
        }
    )


# 0) Prereqs
add(
    "Platform & Security",
    "0",
    "0.1",
    "Prod URL correct (HTTPS, base_url, db, filestore).",
    "Admin",
    "Y",
)
add(
    "Platform & Security",
    "0",
    "0.2",
    "Cron/workers running (mail queue, scheduled actions).",
    "Admin",
    "Y",
)
add(
    "Platform & Security",
    "0",
    "0.3",
    "Companies configured (TBWA/OMC), timezone, currency, fiscal year.",
    "Admin",
    "Y",
)
add(
    "Platform & Security",
    "0",
    "0.4",
    "Users cleaned up (no demo users; internal users active).",
    "Admin",
    "Y",
)

add(
    "Email (Invite-only)",
    "0",
    "0.5",
    "Outgoing email configured (Gmail OAuth or Zoho SMTP).",
    "Admin",
    "Y",
)
add(
    "Email (Invite-only)",
    "0",
    "0.6",
    "Test: send 1 user invite + 1 password reset email successfully.",
    "Admin",
    "Y",
)
add(
    "Email (Invite-only)",
    "0",
    "0.7",
    "Invite-only domain allowlist set (omc.com,tbwa-smp.com).",
    "Admin",
    "N",
)

add(
    "Ask AI",
    "0",
    "0.8",
    "Provider selected (openai or gemini) in Settings.",
    "Admin",
    "Y",
)
add(
    "Ask AI",
    "0",
    "0.9",
    "API keys configured and stored in system params.",
    "Admin",
    "Y",
)
add(
    "Ask AI",
    "0",
    "0.10",
    "Ask AI smoke test returns response; audit trail tables writable.",
    "Admin",
    "Y",
)

# A) Opening Entries Import (Accrual)
add(
    "Accounting",
    "A",
    "A.1",
    "Create AR Clearing (Current Asset, reconcilable).",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "A",
    "A.2",
    "Create AP Clearing (Current Liability, reconcilable).",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "A",
    "A.3",
    "Validate AR totals: invoices + credits + payments match AR aging.",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "A",
    "A.4",
    "Validate AP totals: bills + refunds + payments match AP aging.",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "A",
    "A.5",
    "If multi-currency: update FX rates prior to import.",
    "Accounting Lead",
    "N",
)
add(
    "Accounting",
    "A",
    "A.6",
    "Import open customer invoices + credit notes.",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "A",
    "A.7",
    "Adjust open customer payments (initial balances per partner) if needed.",
    "Accounting Lead",
    "N",
)
add(
    "Accounting",
    "A",
    "A.8",
    "Gate: AR Clearing nets to 0 after imports/adjustments.",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "A",
    "A.9",
    "Import open vendor bills + refunds.",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "A",
    "A.10",
    "Adjust open vendor payments (initial balances per partner) if needed.",
    "Accounting Lead",
    "N",
)
add(
    "Accounting",
    "A",
    "A.11",
    "Gate: AP Clearing nets to 0 after imports/adjustments.",
    "Accounting Lead",
    "Y",
)

# B) Inventory (optional for agency)
add(
    "Inventory (Optional)",
    "B",
    "B.0",
    "Decision: Do we hold inventory? If NO, skip Section B entirely.",
    "Ops/Finance",
    "Y",
)
add(
    "Inventory (Optional)",
    "B",
    "B.1A",
    "Automated valuation: enable automatic stock accounting.",
    "Inventory/Accounting",
    "N",
)
add(
    "Inventory (Optional)",
    "B",
    "B.2A",
    "Automated valuation: set product category costing + valuation.",
    "Inventory/Accounting",
    "N",
)
add(
    "Inventory (Optional)",
    "B",
    "B.3A",
    "Automated valuation: create Inventory Clearing (reconcilable).",
    "Inventory/Accounting",
    "N",
)
add(
    "Inventory (Optional)",
    "B",
    "B.4A",
    "Automated valuation: set inventory clearing on inventory adjustment location.",
    "Inventory/Accounting",
    "N",
)
add(
    "Inventory (Optional)",
    "B",
    "B.5A",
    "Automated valuation: import products with correct category and costs.",
    "Inventory/Accounting",
    "N",
)
add(
    "Inventory (Optional)",
    "B",
    "B.6A",
    "Automated valuation: import initial on-hand quantities; validate adjustment.",
    "Inventory/Accounting",
    "N",
)
add(
    "Inventory (Optional)",
    "B",
    "B.7A",
    "Gate: inventory value equals opening balance.",
    "Inventory/Accounting",
    "N",
)
add(
    "Inventory (Optional)",
    "B",
    "B.1B",
    "Manual valuation: set products with cost.",
    "Inventory",
    "N",
)
add(
    "Inventory (Optional)",
    "B",
    "B.2B",
    "Manual valuation: import initial on-hand quantities; validate adjustment.",
    "Inventory",
    "N",
)
add(
    "Inventory (Optional)",
    "B",
    "B.3B",
    "Gate: balance sheet does not show stock valuation anomalies.",
    "Inventory/Accounting",
    "N",
)

# C) Trial Balance
add(
    "Accounting",
    "C",
    "C.0",
    "Choose bank method: outstanding receipt/payment OR bank clearing.",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "C",
    "C.1A",
    "If outstanding: create Outstanding Receipts + Outstanding Payments (reconcilable).",
    "Accounting Lead",
    "N",
)
add(
    "Accounting",
    "C",
    "C.2A",
    "If outstanding: assign accounts on bank journal (incoming/outgoing).",
    "Accounting Lead",
    "N",
)
add(
    "Accounting",
    "C",
    "C.3A",
    "If outstanding: map bank account to outstanding receipt account.",
    "Accounting Lead",
    "N",
)
add(
    "Accounting",
    "C",
    "C.1B",
    "If bank clearing: create Bank Clearing account (reconcilable).",
    "Accounting Lead",
    "N",
)
add(
    "Accounting",
    "C",
    "C.2B",
    "If bank clearing: map bank account to Bank Clearing.",
    "Accounting Lead",
    "N",
)
add(
    "Accounting",
    "C",
    "C.4",
    "Prepare original Trial Balance (source spreadsheet).",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "C",
    "C.5",
    "Modify TB AR to AR Clearing; exclude open customer payments.",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "C",
    "C.6",
    "Modify TB AP to AP Clearing; exclude open vendor payments.",
    "Accounting Lead",
    "Y",
)
add(
    "Accounting",
    "C",
    "C.7",
    "Modify inventory accounts per valuation decision (Inventory Clearing if automated).",
    "Accounting Lead",
    "N",
)
add(
    "Accounting",
    "C",
    "C.8",
    "Import TB journal entry and post.",
    "Accounting Lead",
    "Y",
)
add("Accounting", "C", "C.9", "Gate: GL AR Clearing = 0.", "Accounting Lead", "Y")
add("Accounting", "C", "C.10", "Gate: GL AP Clearing = 0.", "Accounting Lead", "Y")
add(
    "Accounting",
    "C",
    "C.11",
    "Gate: GL Inventory Clearing = 0 (if used).",
    "Accounting Lead",
    "N",
)

# D) Close + Tax Command Center
add(
    "Close & Tax (IPAI)",
    "D",
    "D.1",
    "Install/upgrade IPAI modules (month_end_closing, bir_tax_compliance, close_orchestration, ask_ai).",
    "Admin",
    "Y",
)
add(
    "Close & Tax (IPAI)",
    "D",
    "D.2",
    "Standardize ownership: Finance Supervisor = Beng (beng.manalo@omc.com).",
    "PM/Finance",
    "Y",
)
add(
    "Close & Tax (IPAI)",
    "D",
    "D.3",
    "Generate CSVs from workbook (data/source.xlsx) and patch role placeholders (no <<MAP:...>>).",
    "PM/Finance",
    "Y",
)
add(
    "Close & Tax (IPAI)",
    "D",
    "D.4",
    "Import Projects, Stages, Tasks Pass A/B, Calendar via JSON-RPC importer.",
    "Admin",
    "Y",
)
add(
    "Close & Tax (IPAI)",
    "D",
    "D.5",
    "Enable and validate stage gates (Prep→Review, Review→Approval, Approval→Done).",
    "PM/Finance",
    "Y",
)
add(
    "Close & Tax (IPAI)",
    "D",
    "D.6",
    "Create Spreadsheet Command Center dashboard (pivots: stage, assignee, overdue, project).",
    "PM/Finance",
    "Y",
)

# E) Invitation-only onboarding (Ops)
add(
    "Onboarding",
    "E",
    "E.1",
    "Invite internal users (invite-only). Verify each can log in and see required apps.",
    "Admin",
    "Y",
)
add(
    "Onboarding",
    "E",
    "E.2",
    "Disable public signup; ensure portal access rules set appropriately.",
    "Admin",
    "Y",
)

# F) Go/No-Go gates
add(
    "Go/No-Go",
    "F",
    "F.1",
    "Run preflight scripts: placeholders=0, Beng user exists, mail server exists.",
    "Admin",
    "Y",
)
add(
    "Go/No-Go",
    "F",
    "F.2",
    "Accounting gates: AR/AP clearing nets to 0; TB posted; bank method configured.",
    "Accounting Lead",
    "Y",
)
add(
    "Go/No-Go",
    "F",
    "F.3",
    "Operational gates: Close/Tax projects visible; dashboard loads; Ask AI responds.",
    "PM/Finance",
    "Y",
)
add(
    "Go/No-Go",
    "F",
    "F.4",
    "Freeze scope for Day 1; publish runbook + escalation path.",
    "Admin",
    "Y",
)

# Workbook build
wb = Workbook()
ws = wb.active
ws.title = "Checklist"

# Header styling
header_fill = PatternFill("solid", fgColor="111111")  # near-black
header_font = Font(bold=True, color="FFD400")  # TBWA yellow
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(style="thin", color="3A3A3A")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(cols)
ws.row_dimensions[1].height = 28
for c, name in enumerate(cols, 1):
    cell = ws.cell(row=1, column=c, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# Write rows
for r in rows:
    ws.append([r[c] for c in cols])

# Column widths
widths = {
    "A": 18,
    "B": 9,
    "C": 10,
    "D": 58,
    "E": 18,
    "F": 28,
    "G": 12,
    "H": 10,
    "I": 28,
    "J": 14,
    "K": 28,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Alignments and borders
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = border
        if cell.column_letter in ("D", "I", "K"):
            cell.alignment = left
        else:
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

# Freeze header
ws.freeze_panes = "A2"

# Data validations
dv_status = DataValidation(
    type="list", formula1='"{}"'.format(",".join(status_opts)), allow_blank=False
)
dv_gate = DataValidation(
    type="list", formula1='"{}"'.format(",".join(gate_opts)), allow_blank=False
)
ws.add_data_validation(dv_status)
ws.add_data_validation(dv_gate)

status_col = cols.index("Status") + 1
gate_col = cols.index("Gate (Y/N)") + 1
dv_status.add(
    f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{ws.max_row}"
)
dv_gate.add(f"{get_column_letter(gate_col)}2:{get_column_letter(gate_col)}{ws.max_row}")

# Conditional formatting on Status
# Done => green fill, Blocked => red fill, In Progress => yellow fill
fill_done = PatternFill("solid", fgColor="1E7E34")  # green
fill_blocked = PatternFill("solid", fgColor="C82333")  # red
fill_progress = PatternFill("solid", fgColor="FFD400")  # yellow
font_dark = Font(color="111111")
font_light = Font(color="FFFFFF")

status_letter = get_column_letter(status_col)
maxr = ws.max_row

ws.conditional_formatting.add(
    f"A2:K{maxr}",
    FormulaRule(formula=[f'${status_letter}2="Done"'], fill=fill_done, font=font_light),
)
ws.conditional_formatting.add(
    f"A2:K{maxr}",
    FormulaRule(
        formula=[f'${status_letter}2="Blocked"'], fill=fill_blocked, font=font_light
    ),
)
ws.conditional_formatting.add(
    f"A2:K{maxr}",
    FormulaRule(
        formula=[f'${status_letter}2="In Progress"'], fill=fill_progress, font=font_dark
    ),
)

# Table
table_ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
tab = Table(displayName="GoLiveChecklist", ref=table_ref)
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(tab)

# Lookups sheet
ws2 = wb.create_sheet("Lookups")
ws2.append(["Status Options"])
for s in status_opts:
    ws2.append([s])
ws2.append([])
ws2.append(["Gate Options"])
for g in gate_opts:
    ws2.append([g])
ws2.column_dimensions["A"].width = 28
for r in ws2.iter_rows():
    for cell in r:
        cell.alignment = left

# Roles sheet (pre-fill Beng mapping)
ws3 = wb.create_sheet("Roles")
ws3.append(["Role", "Default Login/Email"])
roles = [
    ("Finance Supervisor", "beng.manalo@omc.com"),
    ("Senior Finance Manager", "rey.meran@omc.com"),
    ("Finance Director", "khalil.veracruz@omc.com"),
]
for rr in roles:
    ws3.append(list(rr))
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 34
for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws3.freeze_panes = "A2"

# Summary sheet (simple formulas)
ws4 = wb.create_sheet("Summary")
ws4.append(["Metric", "Value"])
metrics = [
    ("Total tasks", f"=COUNTA(Checklist!C2:C{maxr})"),
    ("Done", f'=COUNTIF(Checklist!{status_letter}2:{status_letter}{maxr},"Done")'),
    (
        "In Progress",
        f'=COUNTIF(Checklist!{status_letter}2:{status_letter}{maxr},"In Progress")',
    ),
    (
        "Blocked",
        f'=COUNTIF(Checklist!{status_letter}2:{status_letter}{maxr},"Blocked")',
    ),
    (
        "Not Started",
        f'=COUNTIF(Checklist!{status_letter}2:{status_letter}{maxr},"Not Started")',
    ),
    ("Gates (Y)", f'=COUNTIF(Checklist!H2:H{maxr},"Y")'),
]
for m, v in metrics:
    ws4.append([m, v])

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 20
ws4.freeze_panes = "A2"
for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.border = border
        if cell.row == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        else:
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

# Save workbook
wb.save(out_xlsx)
print(f"Generated: {out_xlsx}")

# Write CSV (Checklist sheet only)
with open(out_csv, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(cols)
    for r in rows:
        w.writerow([r[c] for c in cols])
print(f"Generated: {out_csv}")
