#!/usr/bin/env python3
"""
Plane Finance PPM Template Seeder

Seeds Finance PPM work items from Excel into Plane workspace via API.
Implements rate limiting (60 req/min) and field ownership contract.

Usage:
    export PLANE_API_KEY=<your_api_key>
    export PLANE_WORKSPACE_SLUG=fin-ops
    export FINOPS_XLSX_PATH=/path/to/finance_ppm_tasks.xlsx
    python scripts/plane/seed_finops_from_xlsx.py

Environment Variables:
    PLANE_API_KEY         - Plane API key for authentication (X-API-Key header)
    PLANE_WORKSPACE_SLUG  - Plane workspace slug (default: fin-ops)
    FINOPS_XLSX_PATH      - Path to Excel file with task data (required)
    PLANE_BASE_URL        - Plane API base URL (default: https://api.plane.so)
    DRY_RUN               - Set to "true" for dry-run mode (no API calls)

References:
    - SSOT Template: ssot/plane/templates/finops_month_end.yaml
    - Canonical Seed: addons/ipai/ipai_finance_close_seed/data/06_tasks_month_end.xml
    - PPM Clarity Spec: spec/ppm-clarity-plane-odoo/plan.md
"""

import os
import sys
import time
import json
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


@dataclass
class PlaneConfig:
    """Plane API configuration"""
    api_key: str
    workspace_slug: str
    base_url: str = "https://api.plane.so"
    rate_limit_sleep: float = 1.1  # 1.1s = ~54.5 req/min (safe margin under 60)


class RateLimiter:
    """Simple rate limiter for Plane API (60 req/min)"""
    def __init__(self, sleep_seconds: float = 1.1):
        self.sleep_seconds = sleep_seconds
        self.last_request = 0.0
        self.request_count = 0

    def wait(self):
        """Throttle to respect rate limit"""
        now = time.time()
        elapsed = now - self.last_request
        if elapsed < self.sleep_seconds:
            sleep_time = self.sleep_seconds - elapsed
            time.sleep(sleep_time)
        self.last_request = time.time()
        self.request_count += 1


def _req(cfg: PlaneConfig, method: str, endpoint: str, data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    Make authenticated request to Plane API

    Args:
        cfg: Plane configuration
        method: HTTP method (GET, POST, PATCH, DELETE)
        endpoint: API endpoint path (e.g., /api/v1/workspaces/{slug}/projects/)
        data: Request payload (for POST/PATCH)

    Returns:
        Response JSON as dict

    Raises:
        HTTPError: API request failed
        URLError: Network error
    """
    url = f"{cfg.base_url}{endpoint}"
    headers = {
        "X-API-Key": cfg.api_key,
        "Content-Type": "application/json"
    }

    body = json.dumps(data).encode('utf-8') if data else None
    req = Request(url, data=body, headers=headers, method=method)

    try:
        with urlopen(req) as response:
            response_data = response.read().decode('utf-8')
            return json.loads(response_data) if response_data else {}
    except HTTPError as e:
        error_body = e.read().decode('utf-8')
        logger.error(f"{method} {endpoint} failed: {e.code} {e.reason}")
        logger.error(f"Response: {error_body}")
        raise
    except URLError as e:
        logger.error(f"{method} {endpoint} network error: {e.reason}")
        raise


def ensure_project(cfg: PlaneConfig, limiter: RateLimiter) -> str:
    """
    Get or create Finance PPM project

    Returns:
        Project ID (UUID)
    """
    limiter.wait()
    logger.info(f"Checking for existing project in workspace: {cfg.workspace_slug}")

    # List existing projects
    projects = _req(cfg, "GET", f"/api/v1/workspaces/{cfg.workspace_slug}/projects/")

    # Find Finance PPM project
    target_name = "Finance PPM — Month-end Close"
    for proj in projects:
        if proj.get("name") == target_name:
            logger.info(f"Found existing project: {proj['id']}")
            return str(proj["id"])

    # Create new project
    limiter.wait()
    logger.info(f"Creating new project: {target_name}")
    payload = {
        "name": target_name,
        "identifier": "FINOPS",
        "description": "Month-end close + BIR tax filing execution, owned in Odoo, tracked in Plane.",
        "network": 2  # Secret (only invited)
    }
    project = _req(cfg, "POST", f"/api/v1/workspaces/{cfg.workspace_slug}/projects/", payload)
    logger.info(f"Created project: {project['id']}")
    return str(project["id"])


def ensure_label(cfg: PlaneConfig, limiter: RateLimiter, project_id: str, name: str, color: str) -> str:
    """
    Get or create label in project

    Returns:
        Label ID (UUID)
    """
    limiter.wait()

    # List existing labels
    labels = _req(cfg, "GET", f"/api/v1/workspaces/{cfg.workspace_slug}/projects/{project_id}/labels/")

    # Find existing label
    for label in labels:
        if label.get("name") == name:
            return str(label["id"])

    # Create new label
    limiter.wait()
    logger.info(f"Creating label: {name} ({color})")
    payload = {"name": name, "color": color}
    label = _req(cfg, "POST", f"/api/v1/workspaces/{cfg.workspace_slug}/projects/{project_id}/labels/", payload)
    return str(label["id"])


def create_work_item(
    cfg: PlaneConfig,
    limiter: RateLimiter,
    project_id: str,
    title: str,
    description: str,
    label_ids: List[str],
    estimate: Optional[int] = None
) -> str:
    """
    Create work item in Plane

    Args:
        cfg: Plane configuration
        limiter: Rate limiter
        project_id: Project UUID
        title: Work item title (from Excel "Detailed Monthly Tasks")
        description: Work item description (from task description field)
        label_ids: List of label UUIDs (category + phase)
        estimate: Effort estimate in hours (from "Planned Hours")

    Returns:
        Work item ID (UUID)
    """
    limiter.wait()

    payload = {
        "name": title,
        "description_html": f"<p>{description}</p>",
        "project_id": project_id,
        "label_ids": label_ids
    }

    if estimate is not None:
        payload["estimate_point"] = estimate

    item = _req(cfg, "POST", f"/api/v1/workspaces/{cfg.workspace_slug}/work-items/", payload)
    logger.info(f"Created work item: {title} (ID: {item['id']})")
    return str(item["id"])


def parse_excel_tasks(xlsx_path: str) -> List[Dict[str, Any]]:
    """
    Parse Excel file with Finance PPM tasks

    Expected columns:
        - Task Category: Category label (e.g., "Payroll & Personnel")
        - Detailed Monthly Tasks: Task title/name
        - Phase: Phase label (e.g., "Phase I")
        - Prep By: Employee code who prepares
        - Review By: Employee code who reviews
        - Approve By: Employee code who approves
        - Planned Hours: Effort estimate (default: 16.0)

    Returns:
        List of task dictionaries
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        sys.exit(1)

    logger.info(f"Reading Excel file: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Read header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    logger.info(f"Excel columns: {headers}")

    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        task = dict(zip(headers, row))
        if task.get("Detailed Monthly Tasks"):  # Skip empty rows
            tasks.append(task)

    logger.info(f"Parsed {len(tasks)} tasks from Excel")
    return tasks


def seed_from_canonical_xml() -> List[Dict[str, Any]]:
    """
    Fallback: Parse canonical Odoo XML seed data directly

    Returns:
        List of task dictionaries (same structure as Excel parser)
    """
    import xml.etree.ElementTree as ET

    xml_path = "addons/ipai/ipai_finance_close_seed/data/06_tasks_month_end.xml"
    logger.info(f"Reading canonical XML seed: {xml_path}")

    tree = ET.parse(xml_path)
    root = tree.getroot()

    tasks = []
    for record in root.findall(".//record[@model='project.task']"):
        # Extract fields
        name = record.find("field[@name='name']").text
        description = record.find("field[@name='description']").text

        # Parse description for RACI info
        # Format: "Phase X: ...\nCategory: ...\nPrep by: ... | Review: ... | Approve: ..."
        lines = description.strip().split('\n')
        phase = lines[0].split(':')[0].strip() if lines else "Unknown"
        category = lines[1].replace('Category:', '').strip() if len(lines) > 1 else "Unknown"
        raci = lines[2] if len(lines) > 2 else ""

        # Parse RACI
        prep_by = ""
        review_by = ""
        approve_by = ""
        if '|' in raci:
            parts = raci.split('|')
            for part in parts:
                if 'Prep by:' in part:
                    prep_by = part.replace('Prep by:', '').strip()
                elif 'Review:' in part or 'Review by:' in part:
                    review_by = part.replace('Review:', '').replace('Review by:', '').strip()
                elif 'Approve:' in part or 'Sign-off by:' in part:
                    approve_by = part.replace('Approve:', '').replace('Sign-off by:', '').strip()

        task = {
            "Detailed Monthly Tasks": name,
            "Task Category": category,
            "Phase": phase,
            "Prep By": prep_by,
            "Review By": review_by,
            "Approve By": approve_by,
            "Planned Hours": 16.0  # Default from XML
        }
        tasks.append(task)

    logger.info(f"Parsed {len(tasks)} tasks from canonical XML")
    return tasks


def main():
    """Main seeder execution"""
    # Load configuration
    api_key = os.getenv("PLANE_API_KEY")
    if not api_key:
        logger.error("PLANE_API_KEY environment variable not set")
        sys.exit(1)

    workspace_slug = os.getenv("PLANE_WORKSPACE_SLUG", "fin-ops")
    xlsx_path = os.getenv("FINOPS_XLSX_PATH")
    base_url = os.getenv("PLANE_BASE_URL", "https://api.plane.so")
    dry_run = os.getenv("DRY_RUN", "false").lower() == "true"

    cfg = PlaneConfig(
        api_key=api_key,
        workspace_slug=workspace_slug,
        base_url=base_url
    )

    limiter = RateLimiter(sleep_seconds=cfg.rate_limit_sleep)

    logger.info("=" * 60)
    logger.info("Plane Finance PPM Template Seeder")
    logger.info("=" * 60)
    logger.info(f"Workspace: {workspace_slug}")
    logger.info(f"Base URL: {base_url}")
    logger.info(f"Rate limit: 60 req/min ({cfg.rate_limit_sleep}s sleep)")
    logger.info(f"Dry run: {dry_run}")
    logger.info("=" * 60)

    if dry_run:
        logger.warning("DRY RUN MODE - No API calls will be made")
        return

    # Parse tasks from Excel or fallback to canonical XML
    if xlsx_path and os.path.exists(xlsx_path):
        tasks = parse_excel_tasks(xlsx_path)
    else:
        logger.warning(f"FINOPS_XLSX_PATH not set or file not found: {xlsx_path}")
        logger.info("Falling back to canonical XML seed data")
        tasks = seed_from_canonical_xml()

    # Ensure project exists
    project_id = ensure_project(cfg, limiter)

    # Load label color mapping from SSOT template
    label_colors = {
        # Phase I
        "Payroll & Personnel": "#f56565",
        "Tax & Provisions": "#ed8936",
        "Rent & Leases": "#ecc94b",
        "Accruals & Expenses": "#48bb78",
        "Prior Period Review": "#38b2ac",
        # Phase II
        "Amortization & Corporate": "#4299e1",
        "Corporate Accruals": "#0bc5ea",
        "Insurance": "#9f7aea",
        "Treasury & Other": "#ed64a6",
        "Regional Reporting": "#f687b3",
        # Phase III
        "Client Billings": "#fc8181",
        "WIP/OOP Management": "#f6ad55",
        "AR Aging - WC": "#f6e05e",
        "CA Liquidations": "#68d391",
        "AP Aging - WC": "#4fd1c5",
        "OOP": "#63b3ed",
        "Asset & Lease Entries": "#7f9cf5",
        "Reclassifications": "#b794f4",
        # Phase IV
        "VAT & Taxes": "#f687b3",
        "Accruals & Assets": "#fc8181",
        "AP Aging": "#f6ad55",
        "Expense Reclassification": "#f6e05e",
        "VAT Reporting": "#68d391",
        "Job Transfers": "#4fd1c5",
        "Accruals": "#63b3ed",
        "WIP": "#7f9cf5",
        # Phase V
        "Sign-off": "#b794f4",
        # Phases
        "Phase I": "#f56565",
        "Phase II": "#ed8936",
        "Phase III": "#ecc94b",
        "Phase IV": "#48bb78",
        "Phase V": "#38b2ac",
    }

    # Ensure all labels exist
    label_ids = {}
    unique_labels = set()
    for task in tasks:
        category = task.get("Task Category", "")
        phase = task.get("Phase", "")
        if category:
            unique_labels.add(category)
        if phase:
            unique_labels.add(phase)

    logger.info(f"Creating {len(unique_labels)} unique labels...")
    for label_name in sorted(unique_labels):
        color = label_colors.get(label_name, "#718096")  # Default gray
        label_id = ensure_label(cfg, limiter, project_id, label_name, color)
        label_ids[label_name] = label_id

    # Create work items
    logger.info(f"Creating {len(tasks)} work items...")
    created_count = 0
    for idx, task in enumerate(tasks, 1):
        title = task.get("Detailed Monthly Tasks", "").strip()
        category = task.get("Task Category", "").strip()
        phase = task.get("Phase", "").strip()
        prep_by = task.get("Prep By", "").strip()
        review_by = task.get("Review By", "").strip()
        approve_by = task.get("Approve By", "").strip()
        planned_hours = task.get("Planned Hours", 16.0)

        # Build description with RACI info
        raci_parts = []
        if prep_by:
            raci_parts.append(f"Prep by: {prep_by}")
        if review_by:
            raci_parts.append(f"Review: {review_by}")
        if approve_by:
            raci_parts.append(f"Approve: {approve_by}")

        description = f"{phase}<br>Category: {category}"
        if raci_parts:
            description += f"<br>{' | '.join(raci_parts)}"

        # Collect label IDs
        item_labels = []
        if category and category in label_ids:
            item_labels.append(label_ids[category])
        if phase and phase in label_ids:
            item_labels.append(label_ids[phase])

        # Create work item
        try:
            work_item_id = create_work_item(
                cfg, limiter, project_id,
                title=title,
                description=description,
                label_ids=item_labels,
                estimate=int(planned_hours) if planned_hours else None
            )
            created_count += 1
            logger.info(f"Progress: {created_count}/{len(tasks)} ({idx}/{len(tasks)})")
        except Exception as e:
            logger.error(f"Failed to create work item {idx}: {title}")
            logger.error(f"Error: {e}")

    # Summary
    logger.info("=" * 60)
    logger.info(f"Seeding complete!")
    logger.info(f"Project ID: {project_id}")
    logger.info(f"Labels created: {len(label_ids)}")
    logger.info(f"Work items created: {created_count}/{len(tasks)}")
    logger.info(f"API requests: {limiter.request_count}")
    logger.info(f"Total time: {limiter.request_count * cfg.rate_limit_sleep:.1f}s")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
