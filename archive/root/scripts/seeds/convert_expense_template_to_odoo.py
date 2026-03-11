#!/usr/bin/env python3
"""
Convert Itemized Expense Report Template (Excel) to Odoo hr.expense import CSV.

This script reads the TBWA-style expense template with EXTRA PAGES sheets
and converts them to Odoo-importable CSV format.

Usage:
    python scripts/seeds/convert_expense_template_to_odoo.py INPUT.xlsx [--output OUTPUT.csv]

Template Expected Structure:
- LIQUIDATION FORM sheet: Contains employee name
- EXTRA PAGES sheets: Contains expense line items with columns:
  - Dates | Particulars | Client | CE Number, if chargeable | Meals | Transpo | Misc

Output: CSV file ready for Odoo Expenses import with columns:
- Expense Date, Description, Product, Total, Employee, Paid By, Company
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("Error: pandas required. Install with: pip install pandas")
    sys.exit(1)


# Map template amount buckets -> Odoo Expense Product names
# These products must exist in Odoo before import
CATEGORY_PRODUCT = {
    "Meals": "Meals (Expense)",
    "Transpo": "Transportation (Expense)",
    "Misc": "Miscellaneous (Expense)",
}

# Defaults
DEFAULT_PAID_BY = "Employee (to reimburse)"  # Odoo: payment_mode=own_account
DEFAULT_COMPANY = "InsightPulse AI"  # Match your Odoo company name exactly
DEFAULT_CURRENCY = "PHP"


def norm_str(v) -> str:
    """Normalize value to string, handling None."""
    return "" if v is None else str(v).strip()


def to_date(v) -> date | None:
    """Convert various date formats to date object."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = norm_str(v)
    if not s:
        return None
    # Accept common patterns
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def read_employee_name(wb: openpyxl.Workbook) -> str:
    """Extract employee name from LIQUIDATION FORM sheet."""
    sheet_names = ["LIQUIDATION FORM", "Liquidation Form", "SUMMARY"]
    ws = None
    for name in sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break

    if not ws:
        return ""

    # Scan for "Name:" label and read adjacent cell
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            val = norm_str(cell.value).lower()
            if val in ("name:", "employee name:", "name"):
                # Read cell to the right
                next_val = norm_str(ws.cell(row=cell.row, column=cell.column + 1).value)
                if next_val:
                    return next_val
    return ""


def find_header_row(ws) -> tuple[int | None, dict]:
    """Find the header row containing 'Particulars' and map column positions."""
    header_row = None
    header_map = {}

    for r in range(1, min(ws.max_row + 1, 50)):  # Check first 50 rows
        row_vals = [norm_str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        # Look for "Particulars" as header indicator
        if any(v.lower() == "particulars" for v in row_vals):
            header_row = r
            for c, v in enumerate(row_vals, start=1):
                if v:
                    # Normalize header names
                    header_map[v] = c
            break

    return header_row, header_map


def extract_rows_from_extra_sheet(ws) -> list[dict]:
    """
    Extract expense rows from an EXTRA PAGES sheet.

    Expected columns:
    Dates | Particulars | Client | CE Number, if chargeable | Meals | Transpo | Misc
    """
    header_row, header_map = find_header_row(ws)

    if not header_row:
        return []

    # Map columns (handle variations in naming)
    col_date = header_map.get("Dates") or header_map.get("Date")
    col_part = header_map.get("Particulars") or header_map.get("Description")
    col_client = header_map.get("Client")
    col_ce = (
        header_map.get("CE Number, if chargeable")
        or header_map.get("CE Number")
        or header_map.get("CE#")
    )
    col_meals = header_map.get("Meals")
    col_transpo = header_map.get("Transpo") or header_map.get("Transportation")
    col_misc = header_map.get("Misc") or header_map.get("Miscellaneous")

    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        d = to_date(ws.cell(r, col_date).value) if col_date else None
        particulars = norm_str(ws.cell(r, col_part).value) if col_part else ""
        client = norm_str(ws.cell(r, col_client).value) if col_client else ""
        ce = norm_str(ws.cell(r, col_ce).value) if col_ce else ""

        # Skip empty rows
        if not d and not particulars and not client and not ce:
            continue

        # Process each amount bucket (one expense per non-zero amount)
        buckets = [("Meals", col_meals), ("Transpo", col_transpo), ("Misc", col_misc)]

        for bucket, col in buckets:
            if not col:
                continue
            amt = ws.cell(r, col).value
            if amt is None or amt == "":
                continue
            try:
                amt_f = float(amt)
            except (ValueError, TypeError):
                continue
            if amt_f <= 0:
                continue

            # Build description
            desc = particulars
            suffix_parts = []
            if client:
                suffix_parts.append(f"Client: {client}")
            if ce:
                suffix_parts.append(f"CE: {ce}")
            suffix_parts.append(bucket)

            if suffix_parts:
                suffix = " | ".join(suffix_parts)
                desc = f"{desc} ({suffix})" if desc else suffix

            out.append(
                {
                    "Expense Date": d.isoformat() if d else "",
                    "Description": desc,
                    "Product": CATEGORY_PRODUCT.get(bucket, f"{bucket} (Expense)"),
                    "Total": round(amt_f, 2),
                    "Client": client,
                    "CE Number": ce,
                    "Bucket": bucket,
                }
            )

    return out


def convert_template(input_path: Path, output_path: Path) -> int:
    """
    Convert expense template Excel to Odoo CSV.

    Returns: Number of expense rows converted.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    print(f"Reading: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    # Extract employee name
    employee = read_employee_name(wb)
    if employee:
        print(f"Employee: {employee}")
    else:
        print("Warning: Could not find employee name in template")

    # Extract expense rows from EXTRA PAGES sheets
    rows = []
    extra_sheets = [
        name
        for name in wb.sheetnames
        if re.fullmatch(r"EXTRA PAGES(\s*\(\d+\))?", name, re.IGNORECASE)
    ]

    if not extra_sheets:
        # Fallback: try any sheet that might have expense data
        for name in wb.sheetnames:
            if "liquidation" not in name.lower() and "summary" not in name.lower():
                extra_sheets.append(name)

    print(f"Processing sheets: {extra_sheets}")

    for sheet_name in extra_sheets:
        ws = wb[sheet_name]
        sheet_rows = extract_rows_from_extra_sheet(ws)
        print(f"  - {sheet_name}: {len(sheet_rows)} rows")
        rows.extend(sheet_rows)

    if not rows:
        print("Warning: No expense line items found. Check template structure.")
        # Create empty CSV with headers
        df = pd.DataFrame(
            [],
            columns=[
                "Expense Date",
                "Description",
                "Product",
                "Total",
                "Employee",
                "Paid By",
                "Company",
                "Currency",
            ],
        )
        df.to_csv(output_path, index=False)
        print(f"Wrote empty CSV: {output_path}")
        return 0

    # Build DataFrame
    df = pd.DataFrame(rows)
    df["Employee"] = employee
    df["Paid By"] = DEFAULT_PAID_BY
    df["Company"] = DEFAULT_COMPANY
    df["Currency"] = DEFAULT_CURRENCY

    # Order columns for Odoo import
    cols = [
        "Expense Date",
        "Description",
        "Product",
        "Total",
        "Currency",
        "Employee",
        "Paid By",
        "Company",
        "Client",
        "CE Number",
        "Bucket",
    ]
    df = df[cols]

    # Write CSV
    df.to_csv(output_path, index=False)
    print(f"\nWrote: {output_path} ({len(df)} expense rows)")

    # Show preview
    print("\nPreview (first 10 rows):")
    print(df.head(10).to_string(index=False))

    # Summary by bucket
    print("\nSummary by Category:")
    summary = df.groupby("Bucket")["Total"].agg(["count", "sum"])
    print(summary.to_string())
    print(f"\nTotal: {df['Total'].sum():,.2f} {DEFAULT_CURRENCY}")

    return len(df)


def main():
    parser = argparse.ArgumentParser(
        description="Convert Itemized Expense Report Template to Odoo CSV"
    )
    parser.add_argument("input", type=str, help="Input Excel file (.xlsx)")
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        default=None,
        help="Output CSV file (default: input_odoo_import.csv)",
    )
    parser.add_argument(
        "--company",
        type=str,
        default=DEFAULT_COMPANY,
        help=f"Company name for import (default: {DEFAULT_COMPANY})",
    )
    parser.add_argument(
        "--currency",
        type=str,
        default=DEFAULT_CURRENCY,
        help=f"Currency code (default: {DEFAULT_CURRENCY})",
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = input_path.with_name(f"{input_path.stem}_odoo_import.csv")

    # Update defaults from args
    global DEFAULT_COMPANY, DEFAULT_CURRENCY
    DEFAULT_COMPANY = args.company
    DEFAULT_CURRENCY = args.currency

    try:
        count = convert_template(input_path, output_path)
        print(f"\n{'='*50}")
        print("Next steps in Odoo:")
        print("1. Go to Expenses → My Expenses")
        print("2. Click gear icon → Import records")
        print(f"3. Upload: {output_path}")
        print("4. Map fields and import")
        print("5. Select imported lines → Create Report → Submit")
        return 0
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
