#!/usr/bin/env python3
"""
Generate Odoo-importable Excel workbook from Project Stack seed YAML files.

Creates a single .xlsx file with multiple sheets in the exact import order:
1. partners_customers
2. analytic_accounts
3. products_services
4. projects
5. task_tags
6. task_stages
7. tasks
8. task_dependencies
9. timesheets

Usage:
    python scripts/seeds/generate_project_stack_xlsx.py [--output OUTPUT_FILE]

Dependencies:
    pip install openpyxl pyyaml
"""

import argparse
import os
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    import yaml
except ImportError:
    print("Error: pyyaml is required. Install with: pip install pyyaml")
    sys.exit(1)


# Styling constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def load_yaml(filepath: Path) -> dict:
    """Load a YAML file and return its contents."""
    if not filepath.exists():
        return {}
    with open(filepath, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f) or {}


def style_header(ws, row=1):
    """Apply header styling to the first row."""
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_column_width(ws):
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        length = min(max(length + 2, 12), 50)  # Min 12, max 50
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length


def create_partners_sheet(wb, data: dict):
    """Create partners_customers sheet."""
    ws = wb.create_sheet("partners_customers")

    headers = [
        "id", "name", "company_type", "is_company", "parent_id/id",
        "email", "phone", "street", "city", "zip", "country_id",
        "customer_rank", "supplier_rank", "ref", "function", "comment"
    ]
    ws.append(headers)
    style_header(ws)

    partners = data.get('partners', [])
    for p in partners:
        row = [
            p.get('id', ''),
            p.get('name', ''),
            p.get('company_type', 'company'),
            1 if p.get('is_company', True) else 0,
            p.get('parent_id', ''),
            p.get('email', ''),
            p.get('phone', ''),
            p.get('street', ''),
            p.get('city', ''),
            p.get('zip', ''),
            p.get('country', 'Philippines'),
            p.get('customer_rank', 1),
            p.get('supplier_rank', 0),
            p.get('ref', ''),
            p.get('function', ''),
            p.get('comment', '')
        ]
        ws.append(row)

    auto_column_width(ws)
    return len(partners)


def create_analytic_accounts_sheet(wb, data: dict):
    """Create analytic_accounts sheet."""
    ws = wb.create_sheet("analytic_accounts")

    headers = ["id", "name", "code", "partner_id/id", "active"]
    ws.append(headers)
    style_header(ws)

    accounts = data.get('analytic_accounts', [])
    for a in accounts:
        row = [
            a.get('id', ''),
            a.get('name', ''),
            a.get('code', ''),
            a.get('partner_id', ''),
            1 if a.get('active', True) else 0
        ]
        ws.append(row)

    auto_column_width(ws)
    return len(accounts)


def create_products_sheet(wb, data: dict):
    """Create products_services sheet."""
    ws = wb.create_sheet("products_services")

    headers = [
        "id", "name", "default_code", "detailed_type", "list_price",
        "standard_price", "uom_id", "invoice_policy", "service_type",
        "active", "sale_ok", "description_sale"
    ]
    ws.append(headers)
    style_header(ws)

    products = data.get('products', [])
    for p in products:
        row = [
            p.get('id', ''),
            p.get('name', ''),
            p.get('default_code', ''),
            p.get('type', 'service'),
            p.get('list_price', 0),
            p.get('standard_price', 0),
            p.get('uom_id', 'Hours'),
            p.get('invoice_policy', 'delivery'),
            p.get('service_type', 'timesheet'),
            1 if p.get('active', True) else 0,
            1 if p.get('sale_ok', True) else 0,
            p.get('description_sale', '')
        ]
        ws.append(row)

    auto_column_width(ws)
    return len(products)


def create_projects_sheet(wb, data: dict):
    """Create projects sheet."""
    ws = wb.create_sheet("projects")

    headers = [
        "id", "name", "partner_id/id", "analytic_account_id/id",
        "allow_timesheets", "allow_billable", "timesheet_product_id/id",
        "privacy_visibility", "active", "description", "date_start", "date", "color"
    ]
    ws.append(headers)
    style_header(ws)

    projects = data.get('projects', [])
    for p in projects:
        desc = (p.get('description', '') or '').replace('\n', ' ').strip()
        row = [
            p.get('id', ''),
            p.get('name', ''),
            p.get('partner_id', ''),
            p.get('analytic_account_id', ''),
            1 if p.get('allow_timesheets', True) else 0,
            1 if p.get('allow_billable', False) else 0,
            p.get('timesheet_product_id', ''),
            p.get('privacy_visibility', 'followers'),
            1 if p.get('active', True) else 0,
            desc,
            p.get('date_start', ''),
            p.get('date', ''),
            p.get('color', 0)
        ]
        ws.append(row)

    auto_column_width(ws)
    return len(projects)


def create_tags_sheet(wb, data: dict):
    """Create task_tags sheet."""
    ws = wb.create_sheet("task_tags")

    headers = ["id", "name", "color"]
    ws.append(headers)
    style_header(ws)

    tags = data.get('tags', [])
    for t in tags:
        row = [
            t.get('id', ''),
            t.get('name', ''),
            t.get('color', 0)
        ]
        ws.append(row)

    auto_column_width(ws)
    return len(tags)


def create_stages_sheet(wb, data: dict):
    """Create task_stages sheet."""
    ws = wb.create_sheet("task_stages")

    headers = ["id", "name", "sequence", "fold", "description"]
    ws.append(headers)
    style_header(ws)

    stages = data.get('stages', [])
    for s in stages:
        row = [
            s.get('id', ''),
            s.get('name', ''),
            s.get('sequence', 10),
            1 if s.get('fold', False) else 0,
            s.get('description', '')
        ]
        ws.append(row)

    auto_column_width(ws)
    return len(stages)


def create_tasks_sheet(wb, data: dict):
    """Create tasks sheet."""
    ws = wb.create_sheet("tasks")

    headers = [
        "id", "name", "project_id/id", "partner_id/id", "stage_id/id",
        "tag_ids/id", "parent_id/id", "depend_on_ids/id", "description",
        "planned_hours", "date_deadline", "date_start", "priority", "sequence"
    ]
    ws.append(headers)
    style_header(ws)

    tasks = data.get('tasks', [])
    for t in tasks:
        # Convert tag list to comma-separated string
        tags = t.get('tag_ids', [])
        tags_str = ','.join(tags) if isinstance(tags, list) else str(tags or '')

        # Convert depends_on list to comma-separated string
        deps = t.get('depends_on', [])
        deps_str = ','.join(deps) if isinstance(deps, list) else str(deps or '')

        desc = (t.get('description', '') or '').replace('\n', ' ').strip()

        row = [
            t.get('id', ''),
            t.get('name', ''),
            t.get('project_id', ''),
            t.get('partner_id', ''),
            t.get('stage_id', ''),
            tags_str,
            t.get('parent_id', ''),
            deps_str,
            desc,
            t.get('planned_hours', ''),
            t.get('date_deadline', ''),
            t.get('date_start', ''),
            t.get('priority', '0'),
            t.get('sequence', 10)
        ]
        ws.append(row)

    auto_column_width(ws)
    return len(tasks)


def create_dependencies_sheet(wb, data: dict):
    """Create task_dependencies sheet (blank template)."""
    ws = wb.create_sheet("task_dependencies")

    headers = ["task_id/id", "depend_on_ids/id"]
    ws.append(headers)
    style_header(ws)

    # Add dependencies from tasks if they exist
    dependencies = data.get('dependencies', [])
    count = 0
    for d in dependencies:
        task_id = d.get('task_id', '')
        deps = d.get('depends_on', [])
        for dep in deps:
            ws.append([task_id, dep])
            count += 1

    auto_column_width(ws)
    return count


def create_timesheets_sheet(wb, data: dict):
    """Create timesheets sheet."""
    ws = wb.create_sheet("timesheets")

    headers = [
        "id", "date", "employee_id/id", "project_id/id", "task_id/id",
        "name", "unit_amount", "account_id/id"
    ]
    ws.append(headers)
    style_header(ws)

    timesheets = data.get('timesheets', [])
    for ts in timesheets:
        row = [
            ts.get('id', ''),
            ts.get('date', ''),
            ts.get('employee_id', ''),
            ts.get('project_id', ''),
            ts.get('task_id', ''),
            ts.get('name', ''),
            ts.get('unit_amount', 0),
            ts.get('account_id', '')
        ]
        ws.append(row)

    auto_column_width(ws)
    return len(timesheets)


def create_instructions_sheet(wb):
    """Create an instructions sheet as the first sheet."""
    ws = wb.active
    ws.title = "INSTRUCTIONS"

    instructions = [
        ["Odoo Project Stack Import Workbook"],
        [""],
        ["IMPORT ORDER (must follow this sequence):"],
        ["1. partners_customers - Customers and stakeholders (res.partner)"],
        ["2. analytic_accounts - Analytic accounts for costing (account.analytic.account)"],
        ["3. products_services - Service products for billing (product.product)"],
        ["4. projects - Projects (project.project)"],
        ["5. task_tags - Task tags (project.tags)"],
        ["6. task_stages - Kanban stages (project.task.type)"],
        ["7. tasks - Tasks with hierarchy (project.task)"],
        ["8. task_dependencies - Task dependencies (if using OCA module)"],
        ["9. timesheets - Timesheet entries (account.analytic.line)"],
        [""],
        ["NOTES:"],
        ["- Column headers use Odoo technical field names with /id suffix for relations"],
        ["- Many2one relations: use External ID (e.g., ipai_partner_acme)"],
        ["- Many2many relations: use comma-separated External IDs"],
        ["- Boolean fields: use 1 for True, 0 for False"],
        ["- Dates: use YYYY-MM-DD format"],
        [""],
        ["REQUIRED MODULES:"],
        ["- CE: project, mail, portal, hr_timesheet, sale_timesheet, account"],
        ["- OCA (optional): project_task_dependency, hr_timesheet_sheet"],
        [""],
        ["Generated: " + str(date.today())],
    ]

    for row in instructions:
        ws.append(row)

    # Style the title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'].font = Font(bold=True)
    ws['A14'].font = Font(bold=True)
    ws['A21'].font = Font(bold=True)

    ws.column_dimensions['A'].width = 80


def main():
    parser = argparse.ArgumentParser(
        description='Generate Odoo-importable Excel workbook from Project Stack seeds'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='seeds/workstreams/project_stack/project_stack_import.xlsx',
        help='Output Excel file path'
    )
    args = parser.parse_args()

    # Get project root
    script_dir = Path(__file__).parent.resolve()
    project_root = script_dir.parent.parent

    # Seed directory
    seed_dir = project_root / 'seeds' / 'workstreams' / 'project_stack'

    print("Project Stack Excel Workbook Generator")
    print("=" * 50)

    # Load all YAML data
    print("Loading seed data...")
    partners_data = load_yaml(seed_dir / '10_partners.yaml')
    analytic_data = load_yaml(seed_dir / '20_analytic_accounts.yaml')
    products_data = load_yaml(seed_dir / '30_products.yaml')
    projects_data = load_yaml(seed_dir / '40_projects.yaml')
    tags_data = load_yaml(seed_dir / '50_tags.yaml')
    stages_data = load_yaml(seed_dir / '60_stages.yaml')
    tasks_data = load_yaml(seed_dir / '70_tasks.yaml')
    timesheets_data = load_yaml(seed_dir / '80_timesheets.yaml')

    # Create workbook
    print("Creating Excel workbook...")
    wb = Workbook()

    # Create sheets in order
    create_instructions_sheet(wb)

    counts = {}
    counts['partners'] = create_partners_sheet(wb, partners_data)
    counts['analytic'] = create_analytic_accounts_sheet(wb, analytic_data)
    counts['products'] = create_products_sheet(wb, products_data)
    counts['projects'] = create_projects_sheet(wb, projects_data)
    counts['tags'] = create_tags_sheet(wb, tags_data)
    counts['stages'] = create_stages_sheet(wb, stages_data)
    counts['tasks'] = create_tasks_sheet(wb, tasks_data)
    counts['deps'] = create_dependencies_sheet(wb, tasks_data)
    counts['timesheets'] = create_timesheets_sheet(wb, timesheets_data)

    # Save workbook
    output_path = project_root / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print()
    print("Sheets created:")
    print(f"  - partners_customers: {counts['partners']} records")
    print(f"  - analytic_accounts: {counts['analytic']} records")
    print(f"  - products_services: {counts['products']} records")
    print(f"  - projects: {counts['projects']} records")
    print(f"  - task_tags: {counts['tags']} records")
    print(f"  - task_stages: {counts['stages']} records")
    print(f"  - tasks: {counts['tasks']} records")
    print(f"  - task_dependencies: {counts['deps']} records")
    print(f"  - timesheets: {counts['timesheets']} records")
    print()
    print(f"Workbook saved to: {output_path}")
    print()
    print("Import order in Odoo:")
    print("-" * 50)
    print("1. partners_customers    → res.partner")
    print("2. analytic_accounts     → account.analytic.account")
    print("3. products_services     → product.product")
    print("4. projects              → project.project")
    print("5. task_tags             → project.tags")
    print("6. task_stages           → project.task.type")
    print("7. tasks                 → project.task")
    print("8. task_dependencies     → (if OCA module installed)")
    print("9. timesheets            → account.analytic.line")


if __name__ == '__main__':
    main()
