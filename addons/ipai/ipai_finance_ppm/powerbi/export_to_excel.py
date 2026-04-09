#!/usr/bin/env python3
"""
Finance PPM — Export Power BI views to Excel workbook.

Exports all pbi.* views from PostgreSQL into a single .xlsx file
that can be uploaded directly to Power BI Service (app.powerbi.com).

Usage:
    python export_to_excel.py                          # defaults
    python export_to_excel.py --db odoo --host ipai-odoo-dev-pg.postgres.database.azure.com
    DB_HOST=localhost DB_USER=tbwa python export_to_excel.py

Output:
    powerbi/FinancePPM_OKR_Data.xlsx  (one sheet per view)
"""
import argparse
import os
import sys
from datetime import datetime

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    sys.exit("Missing: pip install psycopg2-binary")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing: pip install openpyxl")


# ── TBWA\SMP theme colors ───────────────────────────────────────────────────
TEAL = "27727B"
GOLD = "FCCE10"
RED = "C1232B"
BG = "FAFAF6"
HEADER_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=9)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=TEAL)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E8E6E0"),
)

VIEWS = [
    "dim_team_member",
    "dim_project",
    "dim_stage",
    "dim_tag",
    "dim_milestone",
    "fact_task",
    "fact_task_assignment",
    "fact_task_tag",
    "agg_team_performance",
    "agg_stage_distribution",
    "agg_milestone_progress",
]


def export(conn, output_path: str):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    for view in VIEWS:
        cur.execute(f"SELECT * FROM pbi.{view}")  # noqa: S608 — view names are hardcoded above
        rows = cur.fetchall()
        columns = [desc[0] for desc in cur.description] if cur.description else []

        # Sheet name max 31 chars
        sheet_name = view[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                val = row[col_name]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER

        # Auto-width columns
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

        # Freeze header
        ws.freeze_panes = "A2"
        # Auto-filter
        if columns:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    cur.close()

    # Add a summary/cover sheet at the front
    cover = wb.create_sheet(title="Finance PPM OKR", index=0)
    cover.cell(row=1, column=1, value="Finance PPM — OKR Dashboard").font = TITLE_FONT
    cover.cell(row=2, column=1, value="TBWA\\SMP · Power BI Data Export").font = Font(
        name="Calibri", size=10, color="7A8A8C"
    )
    cover.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(
        name="Calibri", size=9, color="7A8A8C"
    )
    cover.cell(row=5, column=1, value="Sheets:").font = Font(name="Calibri", size=10, bold=True)
    for i, view in enumerate(VIEWS):
        cover.cell(row=6 + i, column=1, value=f"  {view}").font = DATA_FONT
    cover.cell(row=6 + len(VIEWS) + 1, column=1, value="Instructions:").font = Font(
        name="Calibri", size=10, bold=True
    )
    instructions = [
        "1. Go to app.powerbi.com",
        "2. My Workspace → New → Semantic model → Upload Excel file",
        "3. Upload this .xlsx — each sheet becomes a table",
        "4. Create relationships per FinancePPM_OKR_Model.json",
        "5. Import theme: FinancePPM_OKR_Theme.json",
        "6. Build report pages per FinancePPM_OKR_ReportLayout.json",
    ]
    for i, line in enumerate(instructions):
        cover.cell(row=7 + len(VIEWS) + 1 + i, column=1, value=line).font = DATA_FONT
    cover.column_dimensions["A"].width = 60

    wb.save(output_path)
    return len(VIEWS)


def main():
    parser = argparse.ArgumentParser(description="Export pbi.* views to Excel for Power BI")
    parser.add_argument("--db", default=os.getenv("DB_NAME", "odoo_dev"))
    parser.add_argument("--host", default=os.getenv("DB_HOST", "localhost"))
    parser.add_argument("--port", default=int(os.getenv("DB_PORT", "5432")), type=int)
    parser.add_argument("--user", default=os.getenv("DB_USER", "tbwa"))
    parser.add_argument("--password", default=os.getenv("DB_PASSWORD", ""))
    parser.add_argument(
        "--output",
        default=os.path.join(os.path.dirname(__file__), "FinancePPM_OKR_Data.xlsx"),
    )
    args = parser.parse_args()

    print(f"Connecting to {args.host}:{args.port}/{args.db} as {args.user}...")
    conn = psycopg2.connect(
        host=args.host,
        port=args.port,
        dbname=args.db,
        user=args.user,
        password=args.password or None,
    )
    conn.set_session(readonly=True)

    print(f"Exporting {len(VIEWS)} views...")
    count = export(conn, args.output)
    conn.close()

    size_kb = os.path.getsize(args.output) / 1024
    print(f"✓ Exported {count} sheets → {args.output} ({size_kb:.0f} KB)")
    print()
    print("Next steps:")
    print("  1. Open app.powerbi.com")
    print("  2. My Workspace → New → Semantic model")
    print("  3. Upload this Excel file")
    print("  4. Create report from the dataset")


if __name__ == "__main__":
    main()
